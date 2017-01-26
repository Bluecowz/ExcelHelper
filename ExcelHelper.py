import openpyxl
import datetime
from Settings import Settings
from PopupHelper import Popups

"""
This class handles navigation and manipulation of the excel spreadsheet.
"""


class ExcelHelper:

    def __init__(self):
        self.settings = Settings()
        self._excel = openpyxl.load_workbook(self.settings.active_xl)
        self._sheet = self._excel.active

        if self._new_sheet():
                Popups.full_warning()

    def _new_sheet(self):
        """
        Checks to see if the current sheet is full.
        :return: If the excel sheet is full.
        """
        if self._sheet.cell(row=self.settings.new_sheet_check_Y, column=self.settings.new_sheet_check_X).value is not None:
            return True
        else:
            return False

    def clear_sheet(self):
        """
        Clears the sheet in ZIP storage. Called when copying an excel file so that the contents of the old file
        is not copied into the new sheet
        :return:
        """
        for y in range(3,8):
            for x in range(2,7):
                self._sheet.cell(row=y,column=x).value = None

    def add_time(self):
        """
        Adds the time to the spread sheet. If day is full it will overwrite the last time.
        :return:
        """
        date = datetime.datetime.today().strftime('%m/%d/%Y')
        for y in range(3, 8):
            xl_date_time = datetime.datetime.strptime(str(self._sheet.cell(row=y,column=1).value), '%m/%d/%Y')
            xl_date = xl_date_time.strftime('%m/%d/%Y')
            if xl_date == date:
                for x in range(2, 6):
                    if self._sheet.cell(row=y,column=x).value is None:
                        self._sheet.cell(row=y,column=x).value = datetime.datetime.now().time().strftime("%I:%M %p")
                        self._excel.save(self.settings.active_xl)
                        return
                self._sheet.cell(row=y,column=5).value = datetime.datetime.now().time().strftime("%I:%M %p")
                self._excel.save(self.settings.active_xl)
                return

    def add_desc(self, desc):
        """
        Adds the string to the current string in the description of that day.
        :param desc: Description to add to the current days description.
        :return:
        """
        date = datetime.datetime.today().strftime('%m/%d/%Y')
        for y in range(3, 8):
            xl_date_time = datetime.datetime.strptime(str(self._sheet.cell(row=y,column=1).value), '%m/%d/%Y')
            xl_date = xl_date_time.strftime('%m/%d/%Y')
            if xl_date == date:
                if self._sheet.cell(row=y, column=self.settings.desc_colm).value is not None:
                    self._sheet.cell(row=y, column=self.settings.desc_colm).value = \
                        self._sheet.cell(row=y, column=self.settings.desc_colm).value + " " + desc
                else:
                    self._sheet.cell(row=y, column=self.settings.desc_colm).value = desc
                self._excel.save(self.settings.active_xl)
                return

    def gen_dates(self):
        """
        Generates the next weeks days for a new spreadsheet.
        :return:
        """
        friday = datetime.datetime.now() + datetime.timedelta(ExcelHelper._day_left(datetime.datetime.now(), 'Friday'))
        to_monday = ExcelHelper._day_right(friday, 'Monday')
        date = friday + datetime.timedelta(days=to_monday)
        for y in range(3, 8):
            self._sheet.cell(row=y, column=1, value=date.strftime("%m/%d/%Y"))
            date = date + datetime.timedelta(1)
        self._excel.save(self.settings.active_xl)

    def week_start(self):
        """
        Get the start of the spread sheet week.
        :return: datetime object
        """
        day = self._sheet.cell(row=3, column=1).value
        return day

    def week_end(self):
        """
        Get the end of the spread sheet week.
        :return: datetime object
        """
        day = self._sheet.cell(row=7, column=1).value
        return day

    @staticmethod
    def get_day(date, day):
        """
        Finds the nearest Monday.
        :param date: datetime object
        :param day: Day to search for.
        :return: Number of days to shift
        """
        left = ExcelHelper._day_left(date, day)
        right = ExcelHelper._day_right(date, day)
        if abs(left) < abs(right):
            return left
        else:
            return right

    @staticmethod
    def _day_left(date, day):
        """
        Searches left of the current date.
        :param date: datetime object
        :return: number of days left
        """
        derp = date.strftime("%A")
        if derp == day:
            return 0
        else:
            return -1 + ExcelHelper._day_left(date + datetime.timedelta(days=-1), day)

    @staticmethod
    def _day_right(date, day):
        """
        Searches right of the current date.
        :param date: datetime object
        :return: number of days right
        """
        derp = date.strftime("%A")
        if derp == day:
            return 0
        else:
            return 1 + ExcelHelper._day_right(date + datetime.timedelta(days=1),day)

